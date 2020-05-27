import openpyxl
import pymysql
import pickle

wb = openpyxl.load_workbook(filename="d:\\python\\cb.xlsx")
desList = ['dinner_id', 'dinner_name', 'description', 'start_time', 'end_time', 'reserve_enable', 'reserve_start_hour', 'reserve_end_hour', 'sort_num', 'dinner_status', 'shop_id', 'create_user_name', 'create_user_id', 'create_time', 'update_user_name', 'update_user_id', 'update_time', 'dinner_price']


def getDict():
    db = pymysql.connect(host="47.95.49.67",
                         port=8066,
                         user="sidb",
                         passwd="vFQ8nsab3Ia",
                         db="base")

    cur = db.cursor()
    cur.execute("select group_id,template_shop_id from zys_group limit 99999999")
    res = cur.fetchall()
    cur.execute("select shop_id,group_id from zys_shop_group limit 99999999")

    res2 = cur.fetchall()
    di1 = dict(res)

    res2 = dict(filter(lambda x: not x[0] in di1.values(), res2))

    result = dict(zip(res2.keys(), map(lambda x: di1[x], res2.values())))

    cur.close()
    db.close()
    return result


def getNameTransfer():
    desExel = wb.worksheets[1]
    l1 = []
    l2 = []
    for i in range(1, desExel.max_row + 1):
        l1.append(desExel.cell(i, 1).value)
        l2.append(desExel.cell(i, 2).value)
    desDict = dict(zip(l1, l2))
    return desDict


def getEIdTransfer():
    desExel = wb.worksheets[2]
    l1 = []
    l2 = []
    for i in range(1, desExel.max_row + 1):
        l1.append(desExel.cell(i, 1).value)
        l2.append(desExel.cell(i, 2).value)
    return dict(zip(l1, l2))


desDict = getNameTransfer()
eIdDict = getEIdTransfer()


def getSql(des: list, valueList: list):
    des2 = list(des)
    des2.append("e_dinner_id")
    if not valueList[1] in desDict.keys():
        print(valueList[1] + "不在这里面")
        valueList.append(0)
    else:
        standardDes = desDict[valueList[1]]
        valueList[1] = standardDes
        valueList.append(eIdDict[standardDes])

    valueList[-3] = "2020-02-13 18:12:18"
    valueList[-4] = "0"
    valueList[-5] = "0"
    valueList[-7] = "0"
    valueList[-8] = "0"
    valueStr = ','.join(map(lambda x: "'" + str(x) + "'", valueList))
    sql = "INSERT INTO `zys_dinner_type_test2` (" + ','.join(des2) + ") VALUES (" + valueStr + ");"
    return sql


def getAbondonIds():
    idExel = wb.worksheets[0]
    l = []
    for i in range(1, idExel.max_row + 1):
        l.append(idExel.cell(i, 1).value)
    return l


def getComSql(key, param):
    value1 = desList[1::]
    value2 = desList[1::]
    value2.append("e_dinner_id")
    value1.append("e_dinner_id")
    value2[9] = str(key)
    print(value1)
    sql = "INSERT INTO `zys_dinner_type_test2` (" + ','.join(value1) + ") select " + ','.join(value2) + " from " \
            "zys_dinner_type_test2 where shop_id = " + str(param) + ";"
    return sql


def getAlllist():
    db = pymysql.connect(host="47.95.49.67",
                         port=8066,
                         user="sidb",
                         passwd="vFQ8nsab3Ia",
                         db="food")
    cur = db.cursor()
    cur.execute("select " + ','.join(desList) + " from zys_dinner_type where dinner_status = 1 limit 2000000")
    res = cur.fetchall()
    pickle.dump(res, open("d:\\python\\1.pkl", "wb"))
    cur.close()
    db.close()
    return res


dict1 = getDict()
res = getAlllist()
# res = pickle.load(open("d:\\python\\1.pkl", "rb"))
abondonids = getAbondonIds()
f = open("d:\\python\\1.sql", "w", encoding="utf-8")
for one in res:
    if (one[10] not in dict1.keys()) and (one[0] not in abondonids):
        f.write(getSql(desList, list(one)) + "\n")
for key in dict1.keys():
    f.write(getComSql(key, dict1[key]) + "\n")
f.close()
